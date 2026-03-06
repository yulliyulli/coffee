"""
카페 주문 시스템 - 주문함 기능 추가
"""
from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import func
import yaml
import json
from datetime import datetime
from typing import List, Dict, Optional
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db, create_tables, Order, OrderItem, Cart

app = FastAPI(
    title="카페 주문 시스템",
    description="카페 주문 시스템 백엔드",
    docs_url=None,
    redoc_url=None
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def startup_event():
    """앱 시작 시 데이터베이스 초기화"""
    create_tables()
    print("🚀 카페 주문 시스템이 시작되었습니다")

def load_menu(cafe: str = None):
    """카페별 메뉴 파일 로드"""
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # 카페별 메뉴 파일 시도
    if cafe:
        cafe_menu_path = os.path.join(current_dir, 'config', 'menus', f'{cafe}.yml')
        if os.path.exists(cafe_menu_path):
            with open(cafe_menu_path, 'r', encoding='utf-8') as file:
                return yaml.safe_load(file)

    # 기본 메뉴 파일 (파란만잔)
    menu_path = os.path.join(current_dir, 'config', 'menus', '파란만잔.yml')
    if os.path.exists(menu_path):
        with open(menu_path, 'r', encoding='utf-8') as file:
            return yaml.safe_load(file)

    # 레거시 메뉴 파일
    legacy_path = os.path.join(current_dir, 'config', 'menu.yml')
    with open(legacy_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

# ==================== 주문함 API ====================

@app.post("/api/carts")
def create_cart(cart_data: Dict, db: Session = Depends(get_db)):
    """새 주문함 생성"""
    try:
        # 기존 활성 주문함 비활성화
        db.query(Cart).filter(Cart.is_active == True).update({"is_active": False})

        # 새 주문함 생성
        new_cart = Cart(
            name=cart_data.get("name", "새 주문함"),
            cafe=cart_data.get("cafe", "파란만잔"),
            single_order=cart_data.get("single_order", False),
            is_active=True
        )
        db.add(new_cart)
        db.commit()
        db.refresh(new_cart)

        print(f"🛒 새 주문함 생성: {new_cart.name} ({new_cart.cafe}) - 1인1메뉴: {new_cart.single_order}")

        return {
            "status": "success",
            "cart": {
                "id": new_cart.id,
                "name": new_cart.name,
                "cafe": new_cart.cafe,
                "single_order": new_cart.single_order,
                "is_active": new_cart.is_active,
                "created_at": new_cart.created_at.isoformat()
            }
        }
    except Exception as e:
        db.rollback()
        print(f"주문함 생성 오류: {e}")
        return {"status": "error", "message": str(e)}

@app.get("/api/carts")
def get_carts(db: Session = Depends(get_db)):
    """주문함 목록 조회"""
    try:
        carts = db.query(Cart).order_by(Cart.created_at.desc()).all()
        return {
            "carts": [
                {
                    "id": c.id,
                    "name": c.name,
                    "cafe": c.cafe,
                    "single_order": c.single_order,
                    "is_active": c.is_active,
                    "created_at": c.created_at.isoformat(),
                    "order_count": len(c.orders)
                }
                for c in carts
            ]
        }
    except Exception as e:
        print(f"주문함 목록 조회 오류: {e}")
        return {"carts": []}

@app.get("/api/carts/active")
def get_active_cart(db: Session = Depends(get_db)):
    """활성 주문함 조회"""
    try:
        cart = db.query(Cart).filter(Cart.is_active == True).first()
        if cart:
            return {
                "cart": {
                    "id": cart.id,
                    "name": cart.name,
                    "cafe": cart.cafe,
                    "single_order": cart.single_order,
                    "is_active": cart.is_active,
                    "created_at": cart.created_at.isoformat()
                }
            }
        return {"cart": None}
    except Exception as e:
        print(f"활성 주문함 조회 오류: {e}")
        return {"cart": None}

@app.put("/api/carts/{cart_id}/activate")
def activate_cart(cart_id: int, db: Session = Depends(get_db)):
    """주문함 활성화"""
    try:
        # 모든 주문함 비활성화
        db.query(Cart).update({"is_active": False})

        # 선택한 주문함 활성화
        cart = db.query(Cart).filter(Cart.id == cart_id).first()
        if cart:
            cart.is_active = True
            db.commit()
            return {"status": "success", "message": f"'{cart.name}' 주문함가 활성화되었습니다."}

        return {"status": "error", "message": "주문함를 찾을 수 없습니다."}
    except Exception as e:
        db.rollback()
        return {"status": "error", "message": str(e)}

@app.delete("/api/carts/{cart_id}")
def delete_cart(cart_id: int, db: Session = Depends(get_db)):
    """주문함 삭제"""
    try:
        cart = db.query(Cart).filter(Cart.id == cart_id).first()
        if cart:
            db.delete(cart)
            db.commit()
            return {"status": "success", "message": "주문함가 삭제되었습니다."}
        return {"status": "error", "message": "주문함를 찾을 수 없습니다."}
    except Exception as e:
        db.rollback()
        return {"status": "error", "message": str(e)}

# ==================== 메뉴 API ====================

@app.get("/api/menu")
def get_menu(cafe: Optional[str] = None):
    """메뉴 조회 (카페별)"""
    return load_menu(cafe)

# ==================== 주문 API ====================

@app.get("/api/orders/{customer_name}")
def get_customer_orders(customer_name: str, db: Session = Depends(get_db)):
    """특정 고객의 주문 내역 조회 (활성 주문함 기준)"""
    try:
        # 활성 주문함 찾기
        active_cart = db.query(Cart).filter(Cart.is_active == True).first()

        query = db.query(Order).filter(Order.customer_name == customer_name)
        if active_cart:
            query = query.filter(Order.cart_id == active_cart.id)

        orders = query.order_by(Order.created_at.desc()).all()

        orders_data = []
        for order in orders:
            order_dict = {
                "order_id": order.order_id,
                "customer_name": order.customer_name,
                "timestamp": order.created_at.isoformat(),
                "items": [{"name": item.drink_name, "options": item.options} for item in order.items]
            }
            orders_data.append(order_dict)

        return {"orders": orders_data, "count": len(orders_data)}

    except Exception as e:
        print(f"주문 조회 오류: {e}")
        return {"orders": [], "count": 0}

@app.delete("/api/orders/{customer_name}")
def delete_customer_orders(customer_name: str, db: Session = Depends(get_db)):
    """특정 고객의 모든 주문 삭제"""
    try:
        orders = db.query(Order).filter(Order.customer_name == customer_name).all()

        for order in orders:
            db.delete(order)

        db.commit()
        return {"status": "success", "message": f"{customer_name}님의 주문이 삭제되었습니다."}

    except Exception as e:
        db.rollback()
        return {"status": "error", "message": "주문 삭제에 실패했습니다."}

@app.post("/api/orders/{customer_name}/items")
def add_item_to_order(customer_name: str, item_data: Dict, db: Session = Depends(get_db)):
    """주문 아이템 추가 (활성 주문함에) - 동시성 안전"""
    try:
        # 활성 주문함 찾기
        active_cart = db.query(Cart).filter(Cart.is_active == True).first()
        cart_id = active_cart.id if active_cart else None

        # 1인 1메뉴 제한인 경우: 먼저 주문 생성 시도 (unique constraint로 보호)
        if active_cart and active_cart.single_order:
            # 기존 주문 확인
            existing_order = db.query(Order).filter(
                Order.customer_name == customer_name,
                Order.cart_id == cart_id
            ).first()

            if existing_order:
                print(f"⚠️ 1인1메뉴 제한: {customer_name}님은 이미 주문하셨습니다.")
                raise HTTPException(
                    status_code=400,
                    detail=f"{customer_name}님은 이미 주문하셨습니다. (1인 1메뉴 제한)"
                )

            # 새 주문 생성 (1인 1메뉴)
            order_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:20]
            new_order = Order(
                order_id=order_id,
                customer_name=customer_name,
                cart_id=cart_id
            )
            db.add(new_order)
            db.flush()

            new_item = OrderItem(
                order_id=new_order.id,
                drink_name=item_data["name"],
                options=item_data.get("options", {})
            )
            db.add(new_item)
        else:
            # 일반 모드: 기존 주문에 추가하거나 새로 생성
            existing_order = db.query(Order).filter(
                Order.customer_name == customer_name,
                Order.cart_id == cart_id
            ).order_by(Order.created_at.desc()).first()

            if existing_order:
                new_item = OrderItem(
                    order_id=existing_order.id,
                    drink_name=item_data["name"],
                    options=item_data.get("options", {})
                )
                db.add(new_item)
                existing_order.updated_at = datetime.now()
            else:
                order_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:20]
                new_order = Order(
                    order_id=order_id,
                    customer_name=customer_name,
                    cart_id=cart_id
                )
                db.add(new_order)
                db.flush()

                new_item = OrderItem(
                    order_id=new_order.id,
                    drink_name=item_data["name"],
                    options=item_data.get("options", {})
                )
                db.add(new_item)

        db.commit()
        print(f"✅ 주문 저장: {customer_name} - {item_data['name']}")
        return {"status": "success", "message": "주문이 추가되었습니다."}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        # 동시 주문으로 인한 중복 에러 처리
        if "UNIQUE constraint" in str(e) or "duplicate" in str(e).lower():
            raise HTTPException(
                status_code=400,
                detail=f"{customer_name}님은 이미 주문하셨습니다. (1인 1메뉴 제한)"
            )
        print(f"❌ 주문 저장 실패: {e}")
        return {"status": "error", "message": str(e)}

# ==================== 관리자 API ====================

@app.get("/api/admin/orders")
def get_admin_orders(cart_id: Optional[int] = None, db: Session = Depends(get_db)):
    """관리자용 주문 내역 조회 (주문함 기준)"""
    try:
        # 주문함 ID가 없으면 활성 주문함 사용
        if cart_id is None:
            active_cart = db.query(Cart).filter(Cart.is_active == True).first()
            cart_id = active_cart.id if active_cart else None

        # 해당 주문함의 주문만 조회
        query = db.query(Order)
        if cart_id:
            query = query.filter(Order.cart_id == cart_id)

        orders = query.order_by(Order.created_at.desc()).all()

        orders_data = []
        customers_summary = {}
        total_customers = set()

        for order in orders:
            customer_name = order.customer_name
            total_customers.add(customer_name)

            if customer_name not in customers_summary:
                customers_summary[customer_name] = {
                    "customer_name": customer_name,
                    "order_count": 0,
                    "total_items": 0,
                    "items_list": [],
                    "last_order_date": None,
                    "orders_summary": []
                }

            items_text = []
            for item in order.items:
                options_text = [f"{k}: {v}" for k, v in item.options.items()]
                item_text = f"{item.drink_name} ({', '.join(options_text)})" if options_text else item.drink_name
                items_text.append(item_text)
                customers_summary[customer_name]["items_list"].append(item_text)

            customers_summary[customer_name]["order_count"] += 1
            customers_summary[customer_name]["total_items"] += len(order.items)
            customers_summary[customer_name]["last_order_date"] = order.created_at.strftime("%Y-%m-%d %H:%M")
            customers_summary[customer_name]["orders_summary"].append({
                "order_id": order.order_id,
                "items": items_text,
                "date": order.created_at.strftime("%Y-%m-%d %H:%M")
            })

            orders_data.append({
                "id": order.id,
                "order_id": order.order_id,
                "customer_name": order.customer_name,
                "items_text": " | ".join(items_text),
                "items_count": len(order.items),
                "order_date": order.created_at.strftime("%Y-%m-%d"),
                "order_time": order.created_at.strftime("%H:%M:%S"),
                "timestamp": order.created_at.isoformat()
            })

        customers_list = [
            {
                "customer_name": data["customer_name"],
                "order_count": data["order_count"],
                "total_items": data["total_items"],
                "all_items_summary": " | ".join(data["items_list"]),
                "last_order_date": data["last_order_date"],
                "orders_detail": data["orders_summary"]
            }
            for data in customers_summary.values()
        ]
        customers_list.sort(key=lambda x: x["order_count"], reverse=True)

        return {
            "orders": orders_data,
            "customers_summary": customers_list,
            "total_orders": len(orders_data),
            "total_customers": len(total_customers),
            "total_items": sum(o["items_count"] for o in orders_data)
        }

    except Exception as e:
        print(f"전체 주문 조회 오류: {e}")
        return {"orders": [], "customers_summary": [], "total_orders": 0, "total_customers": 0, "total_items": 0}

@app.get("/api/admin/orders/excel")
def download_orders_excel(cart_id: Optional[int] = None, db: Session = Depends(get_db)):
    """주문 내역 엑셀 다운로드 (주문함 기준)"""
    try:
        if cart_id is None:
            active_cart = db.query(Cart).filter(Cart.is_active == True).first()
            cart_id = active_cart.id if active_cart else None

        query = db.query(Order)
        if cart_id:
            query = query.filter(Order.cart_id == cart_id)

        orders = query.order_by(Order.created_at.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "주문 내역"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")

        headers = ["번호", "고객명", "주문내용", "음료수", "주문일", "주문시간"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        for row_num, order in enumerate(orders, 2):
            items_text = []
            for item in order.items:
                options_text = [f"{k}: {v}" for k, v in item.options.items()]
                item_text = f"{item.drink_name} ({', '.join(options_text)})" if options_text else item.drink_name
                items_text.append(item_text)

            ws.cell(row=row_num, column=1, value=row_num-1)
            ws.cell(row=row_num, column=2, value=order.customer_name)
            ws.cell(row=row_num, column=3, value=" | ".join(items_text))
            ws.cell(row=row_num, column=4, value=len(order.items))
            ws.cell(row=row_num, column=5, value=order.created_at.strftime("%Y-%m-%d"))
            ws.cell(row=row_num, column=6, value=order.created_at.strftime("%H:%M:%S"))

        column_widths = [8, 15, 50, 10, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"주문내역_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            iter([excel_file.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"엑셀 다운로드 오류: {e}")
        raise HTTPException(status_code=500, detail="엑셀 파일 생성에 실패했습니다.")

# ==================== 프론트엔드 정적 파일 서빙 ====================

# 프론트엔드 빌드 디렉토리 경로
current_dir = os.path.dirname(os.path.abspath(__file__))
frontend_dist = os.path.join(current_dir, '..', 'front', 'dist')

# 정적 파일 (assets) 서빙
if os.path.exists(frontend_dist):
    app.mount("/assets", StaticFiles(directory=os.path.join(frontend_dist, "assets")), name="assets")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        """SPA 라우팅 - 모든 경로에서 index.html 반환"""
        # API 경로는 제외
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="Not found")

        file_path = os.path.join(frontend_dist, full_path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)

        # SPA: 모든 경로에서 index.html 반환
        return FileResponse(os.path.join(frontend_dist, "index.html"))

if __name__ == "__main__":
    import uvicorn
    print("🚀 카페 주문 시스템 백엔드를 시작합니다...")
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
